from __future__ import annotations

import base64
from io import BytesIO
from pathlib import Path
from typing import Any

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


NEUTRAL_COLOR = "rgb(82, 54, 36)"


class ReportGenerationError(Exception):
    """Raised when report generation fails with a user-friendly message."""


def _safe_text(value: Any) -> str:
    """Convert any input value to display text, replacing empty values with em dash."""
    if value is None:
        return "—"
    text = str(value).strip()
    return text if text else "—"


def _normalize_numeric(value: Any) -> int | None:
    """
    Convert different Excel numeric formats to integer percent-like value.

    Rules:
    - 0.54 -> 54
    - 54 or 54% -> 54
    - rounded to integer
    - empty/invalid -> None
    """
    if value is None:
        return None

    if isinstance(value, str):
        cleaned = value.strip().replace("%", "").replace(",", ".")
        if not cleaned:
            return None
        try:
            numeric = float(cleaned)
        except ValueError:
            return None
    else:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return None

    if 0 <= numeric <= 1:
        numeric *= 100

    return int(round(numeric))


def _color_for_value(value: int | None) -> str:
    """Choose value color based on requested traffic-light scale."""
    if value is None:
        return NEUTRAL_COLOR
    if 0 <= value <= 30:
        return "rgb(0, 182, 0)"
    if 31 <= value <= 69:
        return "rgb(255, 192, 0)"
    if 70 <= value <= 100:
        return "rgb(225, 0, 0)"
    return NEUTRAL_COLOR


def _read_cell(ws: Any, address: str) -> Any:
    """Read a single worksheet cell with clear error when reference is invalid."""
    try:
        return ws[address].value
    except Exception as exc:  # pragma: no cover - defensive branch
        raise ReportGenerationError(
            f"Не удалось прочитать ячейку {address}. Проверьте структуру Excel-файла."
        ) from exc


def _num_field(ws: Any, address: str) -> dict[str, str]:
    """Build rendering payload for a numeric indicator."""
    normalized = _normalize_numeric(_read_cell(ws, address))
    if normalized is None:
        return {"value": "—", "color": NEUTRAL_COLOR}
    return {"value": str(normalized), "color": _color_for_value(normalized)}


def _load_sheet(file_bytes: bytes) -> Any:
    """Load first worksheet from uploaded xlsx file."""
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except (InvalidFileException, ValueError, OSError, KeyError) as exc:
        raise ReportGenerationError(
            "Не удалось прочитать файл. Загрузите корректный Excel-файл формата .xlsx."
        ) from exc

    if not workbook.sheetnames:
        raise ReportGenerationError("В файле нет листов для чтения.")

    return workbook[workbook.sheetnames[0]]


def generate_report_html(file_bytes: bytes, user_insights: str, template_dir: Path) -> str:
    """
    Build HTML report from uploaded Excel bytes and manual text insights.

    :param file_bytes: Raw uploaded file content.
    :param user_insights: Free text entered by user in UI.
    :param template_dir: Directory that contains Jinja template.
    :return: Rendered standalone HTML report.
    """
    ws = _load_sheet(file_bytes)

    # Декоративный фон (полосы + логотип), экспортированный из background.pdf.
    bg_path = template_dir / "assets" / "report_bg.png"
    try:
        background_base64 = base64.b64encode(bg_path.read_bytes()).decode("ascii")
    except OSError as exc:
        raise ReportGenerationError(
            "Не найден фон templates/assets/report_bg.png. "
            "Положите PNG, экспортированный из topline pre test_ background.pdf."
        ) from exc

    context = {
        "background_data_uri": f"data:image/png;base64,{background_base64}",
        "title_name": _safe_text(_read_cell(ws, "B2")),
        "insights": _safe_text(user_insights),
        "msg1": _safe_text(_read_cell(ws, "A24")),
        "msg2": _safe_text(_read_cell(ws, "A25")),
        "msg3": _safe_text(_read_cell(ws, "A26")),
        "msg4": _safe_text(_read_cell(ws, "A27")),
        "msg5": _safe_text(_read_cell(ws, "A28")),
        "v1": _num_field(ws, "B5"),
        "v2": _num_field(ws, "B16"),
        "v3": _num_field(ws, "B7"),
        "v4": _num_field(ws, "B13"),
        "v5": _num_field(ws, "B11"),
        "v6": _num_field(ws, "B10"),
        "v7": _num_field(ws, "B8"),
        "v8": _num_field(ws, "B7"),
        "v9": _num_field(ws, "B18"),
        "v10": _num_field(ws, "B19"),
        "v11": _num_field(ws, "B24"),
        "v12": _num_field(ws, "B25"),
        "v13": _num_field(ws, "B26"),
        "v15": _num_field(ws, "B27"),
        "v16": _num_field(ws, "B28"),
    }

    env = Environment(
        loader=FileSystemLoader(template_dir),
        autoescape=select_autoescape(["html", "xml"]),
    )

    try:
        template = env.get_template("report_template.html")
    except Exception as exc:
        raise ReportGenerationError("Не найден шаблон templates/report_template.html.") from exc

    return template.render(**context)
